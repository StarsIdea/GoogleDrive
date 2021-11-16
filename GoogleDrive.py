from __future__ import print_function
import sys
import io
import pip
import httplib2
import os
from mimetypes import MimeTypes
import docx
import openpyxl

try:
	from googleapiclient.errors import HttpError
	from apiclient import discovery
	import oauth2client
	from googleapiclient.http import MediaFileUpload, MediaIoBaseDownload
	from oauth2client import file, client, tools

except ImportError:
    print('goole-api-python-client is not installed. Try:')
    print('sudo pip install --upgrade google-api-python-client')
    sys.exit(1)
import sys


class Flag:
    auth_host_name = 'localhost'
    noauth_local_webserver = False
    auth_host_port = [8080, 8090]
    logging_level = 'ERROR'


try:
    import argparse
    flags = Flag()
except ImportError:
    flags = None

SCOPES = 'https://www.googleapis.com/auth/drive'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'GoogleDrive'
dirID = ''
xlsx_name = 'Pakbonnen (1).xlsx'
Shared_Folder = 'Test'
docx_name = '20210411.pdf.txt (1).docx'

def get_credentials():

    home_dir = ''
    # print(home_dir)
    credential_dir = os.path.join(home_dir, '.credentials')
    # print(credential_dir)
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'drive-python-quickstart.json')
    # print(credential_path)
    store = file.Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES)
        flow.user_agent = APPLICATION_NAME
        credentials = tools.run_flow(flow, store, flags)
        # print('Storing credentials to ' + credential_path)
    return credentials


def upload(path, parent_id=None):
    mime = MimeTypes()
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('drive', 'v3', http=http)

    file_metadata = {
        'name': os.path.basename(path)
    }
    if parent_id:
        file_metadata['parents'] = [parent_id]

    media = MediaFileUpload(path,
                            mimetype=mime.guess_type(os.path.basename(path))[0],
                            resumable=True)
    try:
        file = service.files().create(body=file_metadata,
                                  media_body=media,
                                  fields='id, name').execute()
    except HttpError:
        print('corrupted file')
        pass

def listfiles():
    results = service.files().list(fields="nextPageToken, files(id, name,mimeType, shared)").execute()
    items = results.get('files', [])
    xlsx = ''
    if not items:
        print('No files found.')
    else:
        # print('Files:')
        # print('Filename (File ID)')
        for item in items:
            if(item['shared']):
                if item['name'] == Shared_Folder:
                    dirID = item['id']
    results = service.files().list(q="'"+dirID+"' in parents and trashed=false").execute()
    for file_item in results['files']:
        if docx_name in file_item['name']:
            file_id = file_item['id']
        if xlsx_name == file_item['name']:
            xlsx = file_item['id']
    return [file_id, xlsx, dirID]

def delete(fileid):
    service.files().delete(fileId=fileid).execute()


def download(file_id, path=os.getcwd()):
    request = service.files().get_media(fileId=file_id)
    name = service.files().get(fileId=file_id).execute()['name']
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while done is False:
        status, done = downloader.next_chunk()
        # print(int(status.progress() * 100))
    f = open(path + '/' + name, 'wb')
    f.write(fh.getvalue())
    # print('File downloaded at', path)
    f.close()
    # print(file_id)
    # print(name)
    return name

def readFile(filename):
    doc = docx.Document(filename)
    all_paras = doc.paragraphs
    read = False
    result = []
    date = docx_name.split('.')[0]
    for para in all_paras:
        if 'Natuurlijk Vleespakket BV' in para.text:
            read = False
        if read == True:
            array = []
            init_array = para.text.split()
            for i in range(len(init_array)):
                if(i == 0):
                    array.append(date)
                if(i == 1):
                    array.append('')
                if(i == 2):
                    break
                array.append(init_array[i])
                
            Omschrijving = ''
            for i in range(2, len(init_array)):
                Omschrijving += ' '+init_array[i]
            array.append(Omschrijving)
            result.append(array)
        if 'Aantal Eenheid Code Omschrijving' in para.text:
            read = True
    return result

def writeFile(new_data):
    file = xlsx_name
    wb = openpyxl.load_workbook(filename=file)
    ws = wb['Pakbonnen']
    for new_row in new_data:
        row = ws.max_row + 1
        for col, entry in enumerate(new_row, start=1):
            ws.cell(row=row, column=col, value=entry)

    wb.save(file)
    wb.close()

if __name__ == '__main__':
    credentials = get_credentials()
    http = credentials.authorize(httplib2.Http())
    service = discovery.build('drive', 'v3', http=http)

    print('Shared folder name : ')
    Shared_Folder = input()
    print('Insert docx file name : ')
    docx_name = input()
    print('Xlsx file name : ')
    xlsx_name = input()
    print('Inserting ...')
    # method = sys.argv[1]
    file_ids = listfiles()
    docx_file = download(file_ids[0])
    if file_ids[1] != '':
        download(file_ids[1])
        delete(file_ids[1])
    dirID = file_ids[2]
    new_data = readFile(docx_file)
    writeFile(new_data)
    os.remove(docx_file)
    upload(xlsx_name, dirID)
    os.remove(xlsx_name)
    print('Done!')